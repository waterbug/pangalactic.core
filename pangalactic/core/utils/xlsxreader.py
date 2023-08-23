#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Operate on Excel spreadsheets
"""
import argparse
from collections import OrderedDict

from openpyxl import load_workbook

from pangalactic.core import orb


def get_raw_xlsx_data(file_path, clear_empty_rows=True, read_only=False):
    """
    Return the data from an Excel file as a list of lists.

    file_path (str):  name of the file

    clear_empty_rows (bool):  whether to remove empty rows and rows that
        contain only zeros
    read_only (bool):  whether to open in "read only" mode -- this will be
        passed through to openpyxl and puts it into a performance-optimized
        mode for reading large files
    """
    # try:
    book = load_workbook(file_path, read_only=read_only)
    names = book.sheetnames
    # datasets:  a dict that maps sheet names to data contents
    datasets = OrderedDict()
    for name in names:
        datasets[name] = []
        sheet = book[name]
        for row in sheet:
            datasets[name].append([cell.value for cell in row])
    # except IOError:
        # # TODO: use log here
        # orb.log.debug(f'* "{file_path}" not found')
        # return None
    # orb.log.debug('\ncolumn names:  {}'.format(column_names))
    # orb.log.debug('\nfirst row of data:  {}\n'.format(dict(data[0])))
    # for name, value in data[0].items():
        # orb.log.debug('{name:.<40} {value}'.format(name=name, value=value))
    clean_datasets = OrderedDict()
    if clear_empty_rows:
        # orb.log.debug('total rows: {}'.format(len(data))
        for ds_name in datasets:
            clean_datasets[ds_name] = []
            for i, row in enumerate(datasets[ds_name]):
                if not set(row).issubset([0, 0.0, '0', '', None]):
                    clean_datasets[ds_name].append(row)
    else:
        clean_datasets = datasets
    # orb.log.debug('clean data sets:')
    # for name in clean_datasets:
        # orb.log.debug('{}: {} rows'.format(name, len(datasets[name])))
    return clean_datasets


def get_excel_data(file_path, header_pos=None, header_rows=None,
                   key=None, key_type=None, required_col=None,
                   clear_empty_rows=True, read_only=False):
    """
    Return the data from an Excel file as a list of dictionaries.

    Args:
        file_path (str):  name of the file

    Keyword Args:
        header_pos (int):  position of the first header row
        header_rows (int):  number of header row(s)
        key (str):  name of key column (which may be used as a db primary key
            or as a field on which to join the data to another data set)
        key_type (str):  type of key values; options: 'string', 'integer',
            'float';  (default: string)
        required_col (str):  require this column to be populated -- rows in
            which the specified column is empty will be ignored (the main use
            case for this is that if the required column is empty, it signifies
            the end of the data for the current worksheet -- needed for
            worksheets that include empty rows after the data)
        clear_empty_rows (bool):  whether to remove empty rows and rows that
            contain only zeros
        read_only (bool):  whether to open in "read only" mode -- this will be
            passed through to openpyxl and puts it into a performance-optimized
            mode for reading large files
    """
    column_names = []
    data = []
    header_rows = header_rows or 1
    # force key_type to be an acceptable value
    KEY_TYPES = {'string' : str, 'integer' : int, 'float' : float}
    # for LQMS, default key type for "Lab ID" is int
    key_type = key_type or 'integer'
    if key_type not in KEY_TYPES:
        # orb.log.debug('\n  key type "{}" is not allowed;'.format(key_type)
        # orb.log.debug('  key values will be converted to strings.')
        key_type = str
    else:
        key_type = KEY_TYPES[key_type]
    # orb.log.debug('key_type is %s'.format(str(key_type)))
    try:
        book = load_workbook(file_path, read_only=read_only)
        for name, sheet in book.items():
            header_pos, column_names = get_column_names(
                                                sheet,
                                                header_pos=header_pos,
                                                header_rows=header_rows)
            if column_names:
                # orb.log.debug('\n'.join(column_names))
                # now find data
                row = header_pos + header_rows - 2
                while 1:
                    row += 1
                    try:
                        dictified_row = OrderedDict([
                                (colname, sheet.cell_value(row, i))
                                for i, colname in enumerate(column_names)])
                        if ((not required_col) or
                            dictified_row.get(required_col)):
                            # try to coerce key column values to key_type
                            try:
                                if key:
                                    dictified_row[key] = key_type(
                                                        dictified_row[key])
                                data.append(dictified_row)
                            except ValueError:
                                # orb.log.debug(' - key value "{}"'.format(str(
                                        # dictified_row[key])))
                                # orb.log.debug('   cannot be "{}";'.format(
                                        # key_type))
                                # orb.log.debug('   row {} is ignored.\n'.format(
                                        # row))
                                pass
                    except:
                        break
    except IOError:
        # TODO: use log here
        orb.log.debug("\n  Could not find file '{}' -- did you misspell it?\n".format(
              file_path))
    # orb.log.debug("\ncolumn names: {}".format(column_names))
    # orb.log.debug("\nfirst row of data: {}".format(str(dict(data[0]))))
    # for name, value in data[0].items():
        # orb.log.debug('{name:.<40} {value}'.format(name=name, value=value))
    clean_data = []
    if clear_empty_rows:
        # orb.log.debug('total rows: {}'.format(len(data)))
        # orb.log.debug('clearing empty rows ...')
        for i, row in enumerate(data):
            if not set(row.values()).issubset([0, 0.0, '0', '', None]):
                clean_data.append(row)
    else:
        clean_data = data
    # orb.log.debug('remaining rows: {}'.format(len(data)))
    return column_names, clean_data


def get_column_names(sheet, header_pos=None, header_rows=None):
    """
    Return a list of the column names in an Excel file

    @param file_path:  name of the file
    @type  file_path:  L(str)

    @param header_pos:  position of the first header row
    @type  header_pos:  L(int)

    @param header_rows:  number of header row(s)
    @type  header_rows:  L(int)
    """
    found = False
    column_names = []
    header_rows = header_rows or 0
    # if header_pos is not provided, look for first non-empty row ...
    if header_pos is None:
        # orb.log.debug("* header row not specified; searching for first "
              # "non-empty row ...")
        row = -1
        while 1:
            row += 1
            try:
                name = sheet.cell_value(row, 0)
                if name:
                    found = True
                    # orb.log.debug('  - first non-empty row found: {}'.format(
                            # row + 1))
                    column_names = row_to_list(sheet, row)
                    break
                else:
                    continue
            except:
                if found:
                    break
                else:
                    continue
    else:
        row = header_pos - 1
        # orb.log.debug("* header row specified: {}".format(str(header_pos)))
        column_names = row_to_list(sheet, row)
    header_pos = row + 1
    return header_pos, column_names


def row_to_list(sheet, row):
    """
    Return a list containing the values for a specified row in a spreadsheet.
    """
    col = -1
    values = []
    try:
        while 1:
            col += 1
            try:
                values.append(sheet.cell_value(row, col))
            except:
                break
    except:
        orb.log.debug(" - The specified header row is empty.")
    return values


if __name__ == '__main__':
    desc = """Script that reads column names from Excel"""
    parser = argparse.ArgumentParser(description=desc)
    parser.add_argument('file_path',
                        help='path to the Excel file to read')
    parser.add_argument("-c", "--required_column", type=str,
                    help="a column that must be populated [default: None]")
    parser.add_argument("-r", "--header_row_position", type=int, default=0,
                    help="position of first header row [default: 1 (top row)]")
    parser.add_argument("-n", "--number_of_header_rows", type=int, default=1,
                        help="number of header rows [default: 1]")
    parser.add_argument("-d", "--debug", action="store_true",
                        help="debug mode [not yet implemented]")
    parser.add_argument("-t", "--test", action="store_true",
                        help="test mode [not yet implemented]")
    args = parser.parse_args()
    # orb.log.debug(" - supplied args:")
    # for arg, val in vars(args).items():
        # orb.log.debug("   + {}: {}".format(str(arg), str(val)))
    c = None
    r = None
    if args.required_column:
        c = args.required_column
    if args.header_row_position:
        r = args.header_row_position
    get_excel_data(args.file_path, required_col=c, header_pos=r)

